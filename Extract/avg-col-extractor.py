import openpyxl, xlsxwriter, sys, glob

# Week directory (week data from strain-extractor)
weekDir = sys.argv[1]

# Column to get, 1-indexed
col = int(sys.argv[2])

# Output sheet
out = sys.argv[3]

def getWeekAvg(sheets, row, col, cnt):
  y = [] # Average cell values
  for sheet in sheets:
    top = sheet.cell(row, col).value
    if top != None:
      sum = float(top)
      for k in range(row + 1, row + cnt):
        sum += float(sheet.cell(k, col).value)
      y.append(sum / cnt)
    else:
      y.append('')
  return y

# Sort files from weeks, and assume that is the correct time order
dataFiles = glob.glob(f"{weekDir}/*.xlsx")
dataFiles.sort()
weekLabels = [j for j in range(len(dataFiles))]
wbs = [openpyxl.load_workbook(i, read_only=True) for i in dataFiles]
book = xlsxwriter.Workbook(out)
apex = book.add_worksheet("Apex")
apexs = [i["Apex"] for i in wbs]
pp = book.add_worksheet("PP")
pps = [i["PP"] for i in wbs]
smv = book.add_worksheet("SMV")
smvs = [i["SMV"] for i in wbs]
longAxis = book.add_worksheet("Long Axis")
longAxes = [i["Long Axis"] for i in wbs]

for i in zip([apex, pp, smv, longAxis], [apexs, pps, smvs, longAxes], [4, 6, 6, 7]):
  i[0].write_row("A1", ["ID", *weekLabels])

  # j is for output sheet
  # k is for strain
  for j,k in zip(range(1, 7), range(3, 4 + 5 * i[2], i[2]), strict=True):
    i[0].write_row(j, 0, [i[1][0].cell(k, 1).value, *getWeekAvg(i[1], k, col, i[2])])
  i[0].autofit()

book.close()
for i in wbs:
  i.close()